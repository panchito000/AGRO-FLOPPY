#!/usr/bin/env python3
"""Ingesta Excel/PDF/Markdown de Saul → Supabase + conocimiento_index.json (fallback Vercel)."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "Saul" / "Data"
MANIFEST_PATH = DATA_DIR / "documentos_manifest.json"
INDEX_PATH = ROOT / "backend" / "app" / "data" / "conocimiento_index.json"

TIPO_MAP = {
    "siembra": "siembra",
    "temp": "siembra",
    "germin": "siembra",
    "sembr": "siembra",
    "lluvia": "riego",
    "riego": "riego",
    "regar": "riego",
    "hídri": "riego",
    "hidri": "riego",
    "fertil": "fertilizacion",
    "npk": "fertilizacion",
    "plaga": "plagas",
    "insect": "plagas",
    "roya": "plagas",
    "fung": "plagas",
    "herbic": "plagas",
    "maleza": "plagas",
    "cosecha": "cosecha",
    "grano": "cosecha",
    "viento": "plagas",
    "glifosato": "plagas",
    "2,4-d": "plagas",
    "suelo": "siembra",
    "deforest": "siembra",
    "compact": "siembra",
}


def _load_database_url() -> str | None:
    for env_path in (ROOT / ".env", ROOT / ".env.local"):
        if not env_path.exists():
            continue
        for line in env_path.read_text(encoding="utf-8").splitlines():
            if line.startswith("DATABASE_URL="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    return None


def _norm_cultivo(val: str | None) -> str | None:
    if not val:
        return None
    v = val.lower()
    if "soya" in v or "soja" in v:
        return "soya"
    if "maíz" in v or "maiz" in v:
        return "maiz"
    return None


def _infer_tipo_evaluacion(texto: str) -> str | None:
    t = texto.lower()
    for key, tipo in TIPO_MAP.items():
        if key in t:
            return tipo
    return None


def _chunk_text(text: str, size: int = 480, overlap: int = 90) -> list[str]:
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return []
    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = min(len(text), start + size)
        chunk = text[start:end].strip()
        if len(chunk) > 60:
            chunks.append(chunk)
        if end >= len(text):
            break
        start = max(end - overlap, start + 1)
    return chunks


def _limpiar_markdown(texto: str) -> str:
    texto = re.sub(r"^#+\s*", "", texto, flags=re.MULTILINE)
    texto = re.sub(r"\|[^|\n]+\|", " ", texto)
    texto = re.sub(r"!\[[^\]]*\]\([^)]+\)", " ", texto)
    texto = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", texto)
    texto = re.sub(r"---+", " ", texto)
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    return texto


def _agregar_chunks_documento(
    chunks: list[dict],
    *,
    texto: str,
    doc_slug: str,
    doc_titulo: str,
    doc_tipo: str,
    ruta_origen: str,
    fuente_cita: str,
    etiquetas_base: list[str],
    cultivo_default: str | None,
) -> None:
    for piece in _chunk_text(texto):
        lower = piece.lower()
        cultivo = cultivo_default or _norm_cultivo(piece)
        tipo_ev = _infer_tipo_evaluacion(piece)
        etiquetas = list(etiquetas_base)
        if "agricol" in lower:
            etiquetas.append("agricola")
        if "ganader" in lower:
            etiquetas.append("ganadera")
        chunks.append({
            "documento_slug": doc_slug,
            "documento_titulo": doc_titulo,
            "documento_tipo": doc_tipo,
            "ruta_origen": ruta_origen,
            "cultivo": cultivo,
            "tipo_evaluacion": tipo_ev,
            "etiquetas": etiquetas,
            "contenido": piece,
            "fuente_cita": fuente_cita,
        })


def ingest_excel(chunks: list[dict]) -> None:
    try:
        import openpyxl
    except ImportError:
        print("AVISO: openpyxl no instalado, omitiendo Excel.")
        return

    xlsx = DATA_DIR / "Cosecha_Parametros.xlsx"
    if not xlsx.exists():
        print(f"AVISO: No se encontró {xlsx}")
        return

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    doc_slug = "cosecha-parametros-xlsx"

    for sheet in ("Parametros_Cultivo", "Parametros_Producto"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h else "" for h in rows[0]]
        for row in rows[1:]:
            if not row or not any(row):
                continue
            data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            cultivo_raw = str(data.get("Cultivo") or data.get("Producto") or "")
            cultivo = _norm_cultivo(cultivo_raw) if sheet == "Parametros_Cultivo" else None
            param = str(data.get("Parámetro") or data.get("Variable") or "")
            propuesto = data.get("Valor propuesto (hipótesis inicial)")
            encontrado = data.get("Valor encontrado en la investigación")
            fuente = str(data.get("Fuente") or "Cosecha_Parametros.xlsx")
            estado = str(data.get("Estado") or "")
            notas = str(data.get("Notas") or "")

            contenido = (
                f"{cultivo_raw} — {param}: propuesto {propuesto}; "
                f"investigación: {encontrado}. Estado: {estado}. {notas}"
            ).strip()
            tipo_ev = _infer_tipo_evaluacion(f"{param} {notas}")

            chunks.append({
                "documento_slug": doc_slug,
                "documento_titulo": "Cosecha Parámetros (Excel Saul)",
                "documento_tipo": "excel",
                "ruta_origen": str(xlsx.relative_to(ROOT)),
                "cultivo": cultivo,
                "tipo_evaluacion": tipo_ev,
                "etiquetas": [sheet.lower(), estado.lower()] if estado else [sheet.lower()],
                "contenido": contenido[:2000],
                "fuente_cita": fuente[:500],
            })


def ingest_pdf_archivo(
    chunks: list[dict],
    pdf_path: Path,
    meta: dict,
) -> None:
    try:
        from pypdf import PdfReader
    except ImportError:
        print("AVISO: pypdf no instalado, omitiendo PDF.")
        return

    if not pdf_path.exists() or pdf_path.stat().st_size == 0:
        print(f"AVISO: PDF no disponible {pdf_path}")
        return

    reader = PdfReader(str(pdf_path))
    max_pages = meta.get("max_paginas")
    pages = reader.pages[:max_pages] if max_pages else reader.pages
    partes: list[str] = []
    for page in pages:
        text = page.extract_text() or ""
        if text.strip():
            partes.append(text)

    combined = "\n".join(partes)
    if len(combined.strip()) < 100:
        print(f"AVISO: PDF con poco texto extraíble {pdf_path.name}")
        return

    _agregar_chunks_documento(
        chunks,
        texto=combined,
        doc_slug=meta["slug"],
        doc_titulo=meta["titulo"],
        doc_tipo="pdf",
        ruta_origen=str(pdf_path.relative_to(ROOT)),
        fuente_cita=meta["fuente_cita"],
        etiquetas_base=meta.get("etiquetas", []),
        cultivo_default=meta.get("cultivo_default"),
    )
    print(f"  PDF: {pdf_path.name} → {len(partes)} páginas")


def ingest_markdown_archivo(
    chunks: list[dict],
    md_path: Path,
    meta: dict,
) -> None:
    if not md_path.exists():
        print(f"AVISO: Markdown no encontrado {md_path}")
        return
    texto = _limpiar_markdown(md_path.read_text(encoding="utf-8"))
    _agregar_chunks_documento(
        chunks,
        texto=texto,
        doc_slug=meta["slug"],
        doc_titulo=meta["titulo"],
        doc_tipo="markdown",
        ruta_origen=str(md_path.relative_to(ROOT)),
        fuente_cita=meta["fuente_cita"],
        etiquetas_base=meta.get("etiquetas", []),
        cultivo_default=meta.get("cultivo_default"),
    )
    print(f"  MD: {md_path.name}")


def ingest_documentos_manifest(chunks: list[dict]) -> None:
    if not MANIFEST_PATH.exists():
        print(f"AVISO: Sin manifiesto {MANIFEST_PATH}")
        return

    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    print(f"Ingestando {len(manifest)} documentos desde manifiesto…")
    for meta in manifest:
        archivo = DATA_DIR / meta["archivo"]
        if meta.get("tipo") == "markdown":
            ingest_markdown_archivo(chunks, archivo, meta)
        else:
            ingest_pdf_archivo(chunks, archivo, meta)


def ingest_faq(chunks: list[dict]) -> None:
    from app.data.faq_agronomico import FAQ_AGRONOMICO

    for i, faq in enumerate(FAQ_AGRONOMICO):
        chunks.append({
            "documento_slug": "faq-agronomico",
            "documento_titulo": "Preguntas frecuentes Zafra AI",
            "documento_tipo": "faq",
            "ruta_origen": "backend/app/data/faq_agronomico.py",
            "cultivo": faq.get("cultivo"),
            "tipo_evaluacion": faq.get("tipo_evaluacion"),
            "etiquetas": ["faq"] + faq.get("pregunta_patrones", [])[:5],
            "contenido": faq["respuesta"],
            "fuente_cita": faq.get("fuente", "FAQ Zafra AI"),
            "metadata": {
                "documento_titulo": "Preguntas frecuentes Zafra AI",
                "faq_id": i,
                "patrones": faq.get("pregunta_patrones", []),
                "categoria": faq.get("categoria", "informativo"),
            },
        })


def ingest_codigo_base(chunks: list[dict]) -> None:
    """Convierte conocimiento Python existente en chunks indexables."""
    from app.data.conocimiento_agronomico import COSECHA, FERTILIZACION, RIEGO, SIEMBRA
    from app.data.plagas_zona import PLAGAS_POR_CULTIVO

    for cultivo, plagas in PLAGAS_POR_CULTIVO.items():
        for p in plagas:
            chunks.append({
                "documento_slug": "plagas-zona-py",
                "documento_titulo": "Catálogo plagas Santa Cruz (equipo Zafra)",
                "documento_tipo": "codigo",
                "ruta_origen": "backend/app/data/plagas_zona.py",
                "cultivo": cultivo,
                "tipo_evaluacion": "plagas",
                "etiquetas": ["plagas", p["tipo"]],
                "contenido": (
                    f"{p['nombre']}: {p['sintomas']} Monitoreo: {p['monitoreo']} Manejo: {p['manejo']}"
                ),
                "fuente_cita": "Referencias agronómicas Santa Cruz — curado por equipo",
            })

    bloques = [
        ("siembra", SIEMBRA),
        ("riego", RIEGO),
        ("fertilizacion", FERTILIZACION),
        ("cosecha", COSECHA),
    ]
    for tipo, data in bloques:
        for cultivo, info in data.items():
            chunks.append({
                "documento_slug": "conocimiento-agronomico-py",
                "documento_titulo": "Guía agronómica Zafra AI",
                "documento_tipo": "codigo",
                "ruta_origen": "backend/app/data/conocimiento_agronomico.py",
                "cultivo": cultivo,
                "tipo_evaluacion": tipo,
                "etiquetas": [tipo, cultivo],
                "contenido": json.dumps(info, ensure_ascii=False),
                "fuente_cita": "Guía agronómica Zafra — Santa Cruz, Bolivia",
            })


TIPO_DB_MAP = {"markdown": "pdf", "faq": "codigo"}


def persist_supabase(chunks: list[dict], database_url: str) -> None:
    import psycopg2
    from psycopg2.extras import Json

    migration = ROOT / "database" / "migrations" / "002_knowledge_base.sql"
    if migration.exists():
        conn = psycopg2.connect(database_url)
        conn.autocommit = True
        with conn.cursor() as cur:
            cur.execute(migration.read_text(encoding="utf-8"))
        conn.close()

    conn = psycopg2.connect(database_url)
    conn.autocommit = False
    doc_ids: dict[str, int] = {}
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM documento_chunks")
            cur.execute("DELETE FROM documentos")

            for slug in {c["documento_slug"] for c in chunks}:
                sample = next(c for c in chunks if c["documento_slug"] == slug)
                cur.execute(
                    """
                    INSERT INTO documentos (slug, titulo, tipo, ruta_origen, descripcion)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        slug,
                        sample["documento_titulo"],
                        TIPO_DB_MAP.get(sample["documento_tipo"], sample["documento_tipo"]),
                        sample.get("ruta_origen"),
                        f"Ingesta automática — {slug}",
                    ),
                )
                doc_ids[slug] = cur.fetchone()[0]

            for ch in chunks:
                cur.execute(
                    """
                    INSERT INTO documento_chunks
                    (documento_id, cultivo, tipo_evaluacion, etiquetas, contenido, fuente_cita, metadata)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        doc_ids[ch["documento_slug"]],
                        ch.get("cultivo"),
                        ch.get("tipo_evaluacion"),
                        ch.get("etiquetas") or [],
                        ch["contenido"],
                        ch.get("fuente_cita"),
                        Json(ch.get("metadata") or {"documento_titulo": ch["documento_titulo"]}),
                    ),
                )
        conn.commit()
        print(f"Supabase: {len(chunks)} chunks en {len(doc_ids)} documentos.")
    except Exception as exc:
        conn.rollback()
        raise exc
    finally:
        conn.close()


def write_index(chunks: list[dict]) -> None:
    INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 2,
        "total_chunks": len(chunks),
        "chunks": chunks,
    }
    INDEX_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    size_mb = INDEX_PATH.stat().st_size / (1024 * 1024)
    print(f"Índice local: {INDEX_PATH} ({len(chunks)} chunks, {size_mb:.2f} MB)")


def main() -> int:
    sys.path.insert(0, str(ROOT / "backend"))
    chunks: list[dict] = []

    ingest_excel(chunks)
    ingest_documentos_manifest(chunks)
    ingest_faq(chunks)
    ingest_codigo_base(chunks)

    if not chunks:
        print("ERROR: No se generaron chunks.")
        return 1

    write_index(chunks)

    database_url = _load_database_url()
    if database_url:
        try:
            persist_supabase(chunks, database_url)
        except Exception as exc:
            print(f"AVISO Supabase: {exc}")
    else:
        print("AVISO: DATABASE_URL no definida — solo índice local.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
